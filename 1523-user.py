import psutil
import colorama
import os
import wmi
import pyaudio
#import time
import sounddevice as sd
import subprocess
#from termcolor import colored
#import requests
import wave
import keyboard
import openpyxl
import datetime

os.system("cls")
colorama.init()





# Начало записи времени работы
#start = time.time()



# Автор кода 
print(colorama.Fore.MAGENTA + "After: NoFranxx")
print(colorama.Fore.CYAN + "")


# Проверка объёма оперативной памяти
memory_usage = psutil.virtual_memory().total

ram_size = psutil.virtual_memory().total

if ram_size > 12 * 10 ** 9: 
    print(colorama.Fore.GREEN + "Memory_GOOD")
else:
    print(colorama.Fore.RED + "Memory_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("C:/test/error.txt", "a") as file:
     file.write("\n" + "Tester Evgenya Memory_Fail QR-Code материнской платы: " + user_input)

# Отображение кол-ва дисков
c = wmi.WMI()

disk_devices = c.Win32_LogicalDisk()
print(colorama.Fore.WHITE + "Количество накопителей: %s" % len(disk_devices))
if len(disk_devices) > 8 :
    print(colorama.Fore.GREEN + "Storage_GOOD")
else:
    print(colorama.Fore.RED + "Storage_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("C:/test/error.txt", "a") as file:
     file.write("\n" + "Tester Evgenya Storage_Fail QR-Code материнской платы: " + user_input)

#Проверка подключения зарядного устройства
#battery = psutil.sensors_battery()

#if battery.power_plugged:
#    print(colorama.Fore.GREEN + "Зарядное устройство подключено")
#else:
#    print(colorama.Fore.RED + "Зарядное устройство не подключено")
#    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
#    with open("//vdspc/Share2/error.txt", "a") as file:
#     file.write("\n" + "Tester Evgenya Charging_Fail QR-Code материнской платы: " + user_input)
 
# Проверка  микрофона сток
def check_microphone(input_data, frames, time, status):
 if any(input_data > 0):
    if len(input_data) > 3 :
       print(colorama.Fore.GREEN + "Micro_GOOD")
 else:
  print(colorama.Fore.RED + "Micro_Fail")
  user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
  with open("C:/test/error.txt", "a") as file:
   file.write("\n" + "Tester Evgenya Micro_Fail QR-Code материнской платы: " + user_input)

duration = 10
sample_rate = 44100

with sd.InputStream(callback=check_microphone, channels=1, samplerate=sample_rate):
    sd.sleep(duration * 5)
        
 #Проверка Wi-Fi
result = subprocess.run(['netsh', 'wlan', 'show', 'network'], capture_output=True, text=True, encoding='cp866')
networks = result.stdout
count = networks.count('SSID')

# Проверяем количество сетей и выводим соответствующее сообщение
if count > 0:
    print(colorama.Fore.GREEN + "Wi-Fi_Good")
else:
    print(colorama.Fore.RED + "Wi-Fi_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("//vdspc/Share2/error.txt", "a") as file:
     file.write("\n" + "Tester Evgenya Wi-Fi_Fail QR-Code материнской платы: " + user_input)       

# Проверка Ethernet
Ethernet = '//vdspc/Share2/Ethernet.txt'

if os.path.exists(Ethernet):
    print(colorama.Fore.GREEN + 'Ethernet_Good')
else:
   print(colorama.Fore.RED + 'Ethernet_Fail')
   user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
   with open("//vdspc/Share2/error.txt", "a") as file:
    file.write("\n" + "Tester Evgenya Ethernet_Fail QR-Code материнской платы: " + user_input)
 
# Bluetooth
import subprocess
exe_file_path = 'C:/test/test_BT.exe'
result = subprocess.run(exe_file_path, capture_output=True, text=True)

# Проверяем, выдаёт ли файл сообщение "1,2,3" (Это количество сетей bluetooth, которые нашёл файл test_BT.exe)
output = result.stdout.strip()
if output == '1':
    print(colorama.Fore.GREEN + 'Bluetooth_Good')
elif output == '2' or output == '3':
    print(colorama.Fore.GREEN + 'Bluetooth_Good')
else:
   print(colorama.Fore.RED + 'Bluetooth_Bad')
   user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
   with open("C:/test/error.txt", "a") as file:
    file.write("\n" + "Tester Evgenya Bluetooth_Bad QR-Code материнской платы: " + user_input)


# Проверка звука (человеком)    
def play_wave(file_path):
    wf = wave.open(file_path, 'rb')

    p = pyaudio.PyAudio()
    stream = p.open(format=p.get_format_from_width(wf.getsampwidth()),
                    channels=wf.getnchannels(),
                    rate=wf.getframerate(),
                    output=True)
    data = wf.readframes(1024)

    while len(data) > 0:
        stream.write(data)
        data = wf.readframes(1024)

    stream.stop_stream()
    stream.close()
    p.terminate()

file_path = 'C:/test/LRC.WAV'
volume_count = 0

print(colorama.Fore.WHITE + "Нажмите 'U' для воспроизведения звука, нажмите 'Enter' для продолжения.")
print(colorama.Fore.WHITE + "Нажмите 'W' для ввода QR-Code материнской платы, нажмите 'Enter' для продолжения.")

# Ожидание событий клавиатуры
while True:
    if keyboard.is_pressed('U') or keyboard.is_pressed('u'):  # Если нажата 'U'
        play_wave(file_path)
        if volume_count == 0:
            print(colorama.Fore.GREEN + "Volume_GOOD")
            volume_count += 1
        elif volume_count == 1:
            print(colorama.Fore.GREEN + "Volume_GOOD_3.5")
            volume_count = 0
    if keyboard.is_pressed('W') or keyboard.is_pressed('w'):  # Если нажата 'W'
        print(colorama.Fore.RED + "Volume_BAD")
        input(colorama.Fore.BLACK + "")
        user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
        with open("C:/test/error.txt", "w") as file:
         file.write("\n" + "Tester Evgenya Volume_BAD QR-Code материнской платы: " + user_input)
    if keyboard.is_pressed('Enter'):  # Если нажат Enter
        break


# Проверка  микрофона наушники
def check_microphone(input_data, frames, time, status):
 if any(input_data > 0):
    if len(input_data) > 3 :
       print(colorama.Fore.GREEN + "Micro_GOOD_3.5")
 else:
  print(colorama.Fore.RED + "Micro_Fail_3.5")
  user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
  with open("C:/test/error.txt", "a") as file:
   file.write("\n" + "Tester Evgenya Micro_Fail_3.5 QR-Code материнской платы: " + user_input)

duration = 10
sample_rate = 44100

with sd.InputStream(callback=check_microphone, channels=1, samplerate=sample_rate):
    sd.sleep(duration * 5)

input(colorama.Fore.BLACK + "")
 



# Ввод серийного номера
def save_to_excel(file_name):
    try:
        book = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        print(colorama.Fore.RED + "Fail: Файл test excel не обнаружен.")
        user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
        with open("//vdspc/Share2/error.txt", "a") as file:
         file.write(colorama.Fore.WHITE + "\n" + "Tester Evgenya Serial_number_Fail QR-Code материнской платы: " + user_input)
    
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    user_input = input(f"{current_time} Tester Evgenya Serial_number QR-Code материнской платы: ") or "Evgenya"


    try:
        sheet = book['Evgenya']
    except KeyError:
        sheet = book.create_sheet('Evgenya')

    max_row = sheet.max_row
    sheet.cell(row=max_row+1, column=1, value=current_time)
    sheet.cell(row=max_row+1, column=2, value=user_input)

    book.save(file_name)
    print(colorama.Fore.GREEN + "Данные успешно добавлены в лист 'Evgenya' файла", file_name)

file_name = '//vdspc/Share2/test_1523.xlsx'


# Ввод barcode
user_input = input(colorama.Fore.WHITE + "Введите barcode: ")

with open("//vdspc/Share2/barcode.txt", "a") as file1, open("C:/test/barcode.txt", "a") as file2:
    file1.write("\n" + "test_Evgenya: " + user_input)
    file2.write("\n" + "test_Evgenya: " + user_input)
    input("")

print(colorama.Fore.GREEN + "Barcode сохранён в barcode.txt и barcode.txt.")


# Окончание работы программы/запись времени работы
#end = time.time()
#print(colorama.Fore.BLACK + "Execution time of the program is- ", end-start)

#with open("C:/test/test_time.txt", "a") as file:
#    file.write("\n" + "User_test: " + str(end-start))

#print(colorama.Fore.BLACK + "Время работы записано в test_time.txt.")



#Прошивка BIOS
#os.startfile('C:/NLxxPU/FlashMEWinX64.bat')

# Выключение системы
if os.name == 'nt':
    os.system('shutdown /s /t 0')




input("")    

#стабильный вариант на 01.04.2023. (Готовый) (не забудь добавить в "C:/test/" все необходимые файлы: test_time.txt; test.txt; LRC.WAV )

# Windows 11 


#Необходимо добавить 

#Ускорить время работы
