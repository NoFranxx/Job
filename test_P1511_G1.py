import psutil
import colorama
import os
import wmi
import pyaudio
import sounddevice as sd
import subprocess
import wave
import keyboard
import openpyxl
import datetime
import win32api
import numpy as np
os.system("cls")
colorama.init()




# Автор кода 
print(colorama.Fore.MAGENTA + "After: NoFranxx")
print(colorama.Fore.CYAN + "")


#Прошивка audio
#inf_file_path = "C:/nl40/Audio/NL4xPUx_CMedia_1.01/Clevo_MP_Tool_V0803_NLx0PU_PID_025B_EQ2/CM65xx_FWUpdTool_PID_025B/Driver/isousb.inf"
#command = f"pnputil /add-driver {inf_file_path} /install"
#subprocess.run(command, shell=True)

#Запуск прошивки audio
#os.startfile('C:/nl40/Audio/NL4xPUx_CMedia_1.01/Clevo_MP_Tool_V0803_NLx0PU_PID_025B_EQ2/CM65xx_FWUpdTool_PID_01ED/CMUpdate.exe')



# Проверка объёма оперативной памяти
memory_usage = psutil.virtual_memory().total

ram_size = psutil.virtual_memory().total

if ram_size > 12 * 10 ** 9: 
    print(colorama.Fore.GREEN + "Memory_Good")
else:
    print(colorama.Fore.RED + "Memory_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("//vdspc/Share2/error.txt", "a") as file:
     file.write("\n" + "Tester Ksenya Memory_Fail QR-Code материнской платы: " + user_input)



# Отображение кол-ва дисков
c = wmi.WMI()

disk_devices = c.Win32_LogicalDisk()
print("Количество накопителей: %s" % len(disk_devices))
if len(disk_devices) >= 6 :
    print(colorama.Fore.GREEN + "Storage_Good")
else:
    print(colorama.Fore.RED + "Storage_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("//vdspc/Share2/error.txt", "a") as file:
     file.write("\n" + "Tester Ksenya Storage_Fail QR-Code материнской платы: " + user_input)
 
 
 
 # Проверка  микрофона сток
#print(colorama.Fore.WHITE + "Проверка микрофона...")

#duration = 1  # Продолжительность проверки в секундах
#sample_rate = 44100  # Частота дискретизации

#try:
#    audio_data = sd.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1, dtype='float64')
#    sd.wait()
#    
#    if np.max(audio_data) > 0:  # Проверяем наличие сигнала
#        print(colorama.Fore.GREEN + "Micro_Good")
#    else:
#        print(colorama.Fore.RED + "Micro_Fail")#
#
#except Exception as e:
#    print(colorama.Fore.RED + "Micro_Fail")
#    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
#    with open("//vdspc/Share2/error.txt", "a") as file:
#        file.write("\n" + "Tester Ksenya Bluetooth_Fail QR-Code материнской платы: " + user_input)
    
    
    


#Проверка Wi-Fi
result = subprocess.run(['netsh', 'wlan', 'show', 'network'], capture_output=True, text=True, encoding='cp866')
networks = result.stdout
count = networks.count('SSID')

# Проверяем количество сетей и выводим соответствующее сообщение
if count > 0:
    print(colorama.Fore.GREEN + "Wi-Fi_Good")
else:
    print(colorama.Fore.RED + "Wi-Fi_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("//vdspc/Share2/error.txt", "a") as file:
     file.write("\n" + "Tester Ksenya Wi-Fi_Fail QR-Code материнской платы: " + user_input)
     

#Проверка количества видеоадаптеров
c = wmi.WMI()

video_adapters = len(c.Win32_VideoController())

if video_adapters == 2:
    print(colorama.Fore.GREEN + "GPU_Good")
elif video_adapters == 1:
    print(colorama.Fore.RED + "GPU_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("//vdspc/Share2/error.txt", "a") as file:
     file.write("\n" + "Tester Ksenya GPU_Fail QR-Code материнской платы: " + user_input)
     
     

#Проверка клавиатуры
text = input(colorama.Fore.WHITE + "Введите текст qwerty: ")

if text == "qwerty":
    print(colorama.Fore.GREEN + 'keyboard_Good')
else:
   print(colorama.Fore.RED + 'keyboard_Fail')
   user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
   with open("//vdspc/Share2/error.txt", "a") as file:
     file.write("\n" + "Tester Ksenya Ethernet_Fail QR-Code материнской платы: " + user_input)
     




#Проверка подключенных мониторов (для 2)
#monitors = win32api.GetSystemMetrics(80)  # 80 - SM_CMONITORS

#if monitors == 3:
#    print(colorama.Fore.GREEN + "Monitor_Good")
#else:
#    print(colorama.Fore.RED + "Monitor_Fail")
#    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
#    with open("//vdspc/Share2/error.txt", "a") as file:
#     file.write("\n" + "Tester Ksenya Monitor_Fail QR-Code материнской платы: " + user_input)





#Проверка подключения зарядного устройства
battery = psutil.sensors_battery()

if battery.power_plugged:
    print(colorama.Fore.GREEN + "Зарядное устройство подключено")
else:
    print(colorama.Fore.RED + "Зарядное устройство не подключено")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("//vdspc/Share2/error.txt", "a") as file:
     file.write("\n" + "Tester Ksenya Charging_Fail QR-Code материнской платы: " + user_input)





# Проверка Ethernet
Ethernet = '//vdspc/Share2/Ethernet.txt'

if os.path.exists(Ethernet):
    print(colorama.Fore.GREEN + 'Ethernet_Good')
else:
   print(colorama.Fore.RED + 'Ethernet_Fail')
   user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
   with open("//vdspc/Share2/error.txt", "a") as file:
    file.write("\n" + "Tester Ksenya Ethernet_Fail QR-Code материнской платы: " + user_input)




 
# Bluetooth
exe_file_path = 'C:/test/test_BT.exe'
result = subprocess.run(exe_file_path, capture_output=True, text=True)

# Проверяем, выдаёт ли файл сообщение "1,2,3" (Это количество сетей bluetooth, которые нашёл файл test_BT.exe)
output = result.stdout.strip()
if output == '1':
    print(colorama.Fore.GREEN + 'Bluetooth_Good')
elif output == '2' or output == '3':
    print(colorama.Fore.GREEN + 'Bluetooth_Good')
else:
   print(colorama.Fore.RED + 'Bluetooth_Fail')
   user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
   with open("//vdspc/Share2/error.txt", "a") as file:
    file.write("\n" + "Tester Ksenya Bluetooth_Fail QR-Code материнской платы: " + user_input)





# Проверка звука (человеком)    
def play_wave(file_path):
    wf = wave.open(file_path, 'rb')

    p = pyaudio.PyAudio()
    stream = p.open(format=p.get_format_from_width(wf.getsampwidth()),
                    channels=wf.getnchannels(),
                    rate=wf.getframerate(),
                    output=True)
    data = wf.readframes(1024)

    while len(data) > 0:
        stream.write(data)
        data = wf.readframes(1024)

    stream.stop_stream()
    stream.close()
    p.terminate()

file_path = 'C:/test/LRC.WAV'
volume_count = 0

print(colorama.Fore.WHITE + "Нажмите 'U' для воспроизведения звука, нажмите 'Enter' для продолжения.")
print(colorama.Fore.WHITE + "Нажмите 'W' для ввода QR-Code материнской платы, нажмите 'Enter' для продолжения.")

# Ожидание событий клавиатуры
while True:
    if keyboard.is_pressed('U') or keyboard.is_pressed('u'):  # Если нажата 'U'
        play_wave(file_path)
        if volume_count == 0:
            print(colorama.Fore.GREEN + "Volume_Good")
            volume_count += 1
        elif volume_count == 1:
            print(colorama.Fore.GREEN + "Volume_Good_3.5")
            volume_count = 0
    if keyboard.is_pressed('W') or keyboard.is_pressed('w'):  # Если нажата 'W'
        print(colorama.Fore.RED + "Volume_Fail")
        input(colorama.Fore.BLACK + "")
        user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
        with open("//vdspc/Share2/error.txt", "w") as file:
         file.write("\n" + "Tester Ksenya Volume_Fail QR-Code материнской платы: " + user_input)
    if keyboard.is_pressed('Enter'):  # Если нажат Enter
        break


input(colorama.Fore.BLACK + "")




# Проверка  микрофона наушники
print(colorama.Fore.WHITE + "Проверка микрофона...")

duration = 1  # Продолжительность проверки в секундах
sample_rate = 44100  # Частота дискретизации

try:
    audio_data = sd.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1, dtype='float64')
    sd.wait()
    
    if np.max(audio_data) > 0:  # Проверяем наличие сигнала
        print(colorama.Fore.GREEN + "Micro_Good")
    else:
        print(colorama.Fore.RED + "Micro_Fail")

except Exception as e:
    print(colorama.Fore.RED + "Micro_Fail")
    user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
    with open("//vdspc/Share2/error.txt", "a") as file:
        file.write("\n" + "Tester Ksenya Bluetooth_Fail QR-Code материнской платы: " + user_input)
print(colorama.Fore.BLACK + "")
print(colorama.Fore.WHITE + "")



# Ввод серийного номера
def save_to_excel(file_name):
    try:
        book = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        print(colorama.Fore.RED + "Fail: Файл test excel не обнаружен.")
        user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
        with open("//vdspc/Share2/error.txt", "a") as file:
         file.write(colorama.Fore.WHITE + "\n" + "Tester Ksenya Serial_number_Fail QR-Code материнской платы: " + user_input)
    
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    user_input = input(f"{current_time} Tester Ksenya Serial_number QR-Code материнской платы: ") or "Ksenya"


    try:
        sheet = book['Ksenya']
    except KeyError:
        sheet = book.create_sheet('Ksenya')

    max_row = sheet.max_row
    sheet.cell(row=max_row+1, column=1, value=current_time)
    sheet.cell(row=max_row+1, column=2, value=user_input)

    book.save(file_name)
    print(colorama.Fore.GREEN + "Данные успешно добавлены в лист 'Ksenya' файла", file_name)

file_name = '//vdspc/Share2/P1511SD.xlsx'


save_to_excel(file_name)





# Ввод barcode
#def save_to_excel_1(file_name_1):
#    try:
#        book = openpyxl.load_workbook(file_name_1)
#    except FileNotFoundError:
#        print(colorama.Fore.RED + "Fail: Файл test excel не обнаружен.")
#        user_input = input(colorama.Fore.WHITE + "Введите QR-Code материнской платы: ")
#        with open("//vdspc/Share2/error.txt", "a") as file:
#         file.write("\n" + "Tester Ksenya Barcode_Fail QR-Code материнской платы: " + user_input_1)
#    
#    current_time_1 = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#    user_input_1 = input(colorama.Fore.WHITE + f"{current_time_1} Tester Ksenya Barcode QR-Code материнской платы: ") or "Ksenya"


#    try:
#        sheet = book['barcode']
#    except KeyError:
#        sheet = book.create_sheet('barcode')
#
#    max_row = sheet.max_row
#    sheet.cell(row=max_row+1, column=1, value=current_time_1)
#    sheet.cell(row=max_row+1, column=2, value=user_input_1)
#
#    book.save(file_name_1)
#    print(colorama.Fore.GREEN + "Данные успешно добавлены в лист 'barcode' файла", file_name_1)

# = '//vdspc/Share2/test.xlsx'
#save_to_excel_1(file_name_1)


# Выключение системы
if os.name == 'nt':
    os.system('shutdown /s /t 0')

input("")    
  

#не стабильный вариант на 08.04.2023

# Windows 11 P1511_G1


#Необходимо добавить отпечаток пальца, мемтест на 30 секунд, бёрн на 30 секунд, проверку по каждой usb, скорость вентиляторов, 

#Ускорить время работы

#